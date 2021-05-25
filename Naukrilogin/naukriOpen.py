from selenium import webdriver
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import workbook
import time

class OpenNaukri(object):
    def __init__(self,driver='driver'):
        self.driver=webdriver.Chrome(executable_path='C:\\Users\\rakes\\Desktop\\chromedriver.exe')
        self.driver.maximize_window()
        self.driver.get_screenshot_as_file('E:\\Python_3X\\workspace_python\\seleniumPratice\\Naukrilogin\\ScreenShots\\withOutOpenURL.png')
        wb = openpyxl.load_workbook('sample.xlsx')
        ws1 = wb.active
        self.urlofsite = ws1.cell(1, 1).value
        self.username = ws1.cell(1, 2).value
        self.password = ws1.cell(1, 3).value


    def naukriOpen(self):
        self.driver.get(self.urlofsite)
        self.driver.implicitly_wait(3)
        self.driver.get_screenshot_as_file(
            'E:\\Python_3X\\workspace_python\\seleniumPratice\\Naukrilogin\\ScreenShots\\afterOpenURL.png')
        currentwindowhandle=self.driver.current_window_handle
        print(currentwindowhandle)
        allhandels=self.driver.window_handles
        print(len(allhandels))
        print(allhandels)
        # for i in range(0, len(allhandels)):
        #     self.driver.switch_to_window(self.driver.window_handles[i])
        #     time.sleep(3)
        #     print('current open site is' +str(self.driver.title))
        #     # self.driver.close()

        self.driver.switch_to_window(currentwindowhandle)
        login=self.driver.find_element_by_css_selector('#login_Layer')
        actions1=ActionChains(self.driver)
        loginaction=actions1.move_to_element(login)
        loginaction.perform()
        more=self.driver.find_element_by_xpath("//div[contains(text(), 'More')]")
        moreaction=actions1.move_to_element(more)
        moreaction.perform()
        time.sleep(2)
        naukriblog=self.driver.find_element_by_xpath("//ul[@class='group']//a[contains(text(), 'Naukri Blog')]")
        naukriblog.click()
        time.sleep(3)
        allhandels2=self.driver.window_handles
        for i in allhandels2:
            if i not in allhandels:
                self.driver.switch_to_window(i)
                print(self.driver.title)
        self.driver.get_screenshot_as_file('more.png')
        self.driver.execute_script('window.scrollTo(0,document.body.scrollHeight)')

        time.sleep(10)
        self.driver.quit()


ON=OpenNaukri()
ON.naukriOpen()





