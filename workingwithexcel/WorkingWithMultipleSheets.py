from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl

wb=Workbook()
print(wb.active.title)
ws1=wb.active
ws1.title='Replaced name'
ws2=wb.create_sheet("School details")
sheets1=wb.sheetnames
ws3=wb.create_sheet("Created by system")
sheets2=wb.sheetnames
wb.remove(ws3)
sheets3=wb.sheetnames
ws1['A1']='Test details'
data2=[(1,2,3),(4,5,6)]
for j in data2:
    ws2.append(j)

data=((1,2,3),
      (4,5,6),
      ('Rakesh','Kannayyagari','Bodagattu'),
      ('Vanitha','Vanaparthi','Hyderabad'),
      ('Conduent','Kokapet','Gandipet'))

for i in data:
    ws1.append(i)
    ws2.append(i)

ws1.append(['Added by append'])

for k in sheets1:
    ws1.append([k])

ws1.append(["adding sheet"])

for l in sheets2:
    ws1.append([l])

ws1.append(["Sheet removed"])
ws1['A1'].fill=PatternFill("solid",fgColor='71FF33')
ws2['A1'].fill=PatternFill("solid",fgColor='71FF33')
for m in sheets3:
    ws1.append([m])

print(type(sheets1))
print(sheets1)
print(sheets2)
print(sheets3)
wb.save('test.xlsx')
# wb.active=1
# print(wb.active.title)
