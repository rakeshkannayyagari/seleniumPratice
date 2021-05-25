import openpyxl

wb=openpyxl.load_workbook('designation.xlsx')
sh1=wb.active
print(sh1.title)
minrow=sh1.min_row
mincol=sh1.min_column
maxrow=sh1.max_row
maxcol=sh1.max_column
wb.create_sheet("location")
print(wb.active.title)


for title in sh1.iter_rows(min_row=1,max_row=1,min_col=1,max_col=maxcol,values_only='True'):
    print(title)

for name in sh1.iter_cols(min_row=2,max_row=maxrow,min_col=1,max_col=maxcol,values_only='True'):
    print(name)

wb.save("designation.xlsx")

import openpyxl
from openpyxl import Workbook
from openpyxl import  worksheet
wb=Workbook()
wb.active.title="Employe details"
wb.create_sheet("Education details")
wb.create_sheet("Experiance")
print(wb.sheetnames)
ws1=wb.active
ws1['A1']='employee details'
ws2=wb.active=1
ws3=wb['Experiance']
ws2['A1']='Sheet2'
ws3['A1']='Sheet3'

wb.save("multiplesheets.xlsx")

