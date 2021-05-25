import openpyxl
from openpyxl.workbook import Workbook
file=openpyxl.load_workbook('TestData.xlsx')
#sheet names printed
sheets=file.sheetnames
print(sheets)
#reading data from sheet Method 1
sheet1=file['Names']
data1=sheet1['B2'].value
print(data1)
#method 2
data2=file['Marks']['A2'].value
print(data2)
#Method 3
data3=sheet1.cell(1,1).value#row, column
print(data3)
#*****************************************************
from openpyxl import Workbook
import openpyxl

class Exceloperations():
    def readvalue(self):
        wb=openpyxl.load_workbook('TestData.xlsx')
        print('Sheet names before modification'+str(wb.sheetnames))
        wb['Names'].title='Student Names'
        wb['Marks'].title='Student Marks'
        wb['Location'].title='Branch'
        print('Sheet names after modification' + str(wb.sheetnames))
        sheet1=wb['Student Names']
        row=sheet1.max_row
        column=sheet1.max_column
        print("number of rows "+str(row))
        print("number of columns " + str(column))
        for i in range(1,row+1):
            for j in range(1,column+1):
                value=sheet1.cell(i,j).value
                print("row "+str(i)+' column '+ str(j)+" "+ str(value))

        #writing data ti excel.
        for i in range(row+1,10):
            for j in range(1,10):
                sheet1.cell(i,j).value=i+j
        wb.save('TestData2.xlsx')


E1=Exceloperations()
E1.readvalue()
#***************************
from openpyxl import load_workbook
from openpyxl import Workbook

wb=Workbook()
sheet1name=wb.active.title
wb.active.title='Test sheet name'
print(sheet1name)
print(wb.active.title)
sheet1=wb['Test sheet name']
sheet1['A1']='Name'
sheet1['B1']='Designation'
sheet1['A2']='Rakesh'
sheet1['B2']='IT Test eng 2'
sheet1['A3']='Vanitha'
sheet1['B3']='Analyst'
sheet1['A4']='Vanaparthi'
sheet1['B4']='Member'

wb.save('designation.xlsx')
